# program for Sheet Automation & Chart Plotting

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

# cell = sheet['a1'] for accessing cells in the sheet in not a good way
# cell = sheet.cell(1, 1)
cell1 = sheet.cell(1, 4)
cell1.value = 'Corrected price'

for row in range(2, sheet.max_row + 1):  # sheet.max_row gives maximum no. of rows in that sheet
    cell = sheet.cell(row, 3)
    # print(cell.value)
    corrected_value = cell.value * 0.9
    corrected_value_cell = sheet.cell(row, 4)
    corrected_value_cell.value = corrected_value

values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

# add_chart adds chart to sheet are chart object & leftmost point in excel sheet cell

wb.save('tranasactions2.xlsx')
